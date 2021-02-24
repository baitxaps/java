from openpyxl import Workbook,load_workbook
import os

#cmd-> pip install openpyxl

def create_sheet():
    wb=Workbook()
    sheet=wb.active 
    print(sheet.title)
    #sheet.title="data"

    sheet["b9"]="girl"
    sheet["c9"]="171,48,89"
    sheet.append(["rachel","180","60"])
    wb.save("excel_test.xlsx")

def load_file():
    # path = os.path.dirname(__file__)
    # path = os.path.dirname(path)+"\\excel_test.xlsx"
    # print(path)
    wb=load_workbook("excel_test.xlsx")
    print(wb.sheetnames)
    print(wb.get_sheet_names())
    sheet=wb.get_sheet_by_name("Sheet")
    print(sheet["b9"])
    print(sheet["b9"].value)

    for col in sheet.columns:
        for cell in col:
            print(cell.value,end=",")
        print()    

    for col in sheet.iter_cols(min_col=5,max_col=10,min_row=3,max_row=20):
        for cell in col:
            print(cell.value,end=",")
        print()    

    # for row in sheet.iter_rows(min_row=5,max_row=10,max_col=9):
    #     for cell in row:
    #         print(cell.value,end=",")
    #     print()    

    # for row in sheet:
    #     print(row)
    #     for cell in row:
    #         print(cell.value,end=",")
    #     print()    


from openpyxl.styles import Font,colors,Alignment
def test_sheet_attributes():
    wb=load_workbook("excel_test.xlsx")
    sheet=wb.get_sheet_by_name("Sheet")

    font=Font(name="宋体",size=20,italic=True,color=colors.BLUE)
    alignment=Alignment(vertical="center",horizontal="center")
    height = 40
    width =30

    sheet.row_dimensions[2].height=height
    sheet.column_dimensions['C'].width=width

    sheet["B5"].alignment= alignment
    sheet["B5"].font = font
    wb.save("excel_test.xlsx")

#create_sheet()
#load_file()
test_sheet_attributes()