from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText #mail 正文
from email.header import Header # mail 

wb = load_workbook("gzb.xlsx",data_only=True)
sheet=wb.active

#login
smtp_obj = smtplib.SMTP_SSL("smtp.qq.com",465)
smtp_obj.login("hnchenhairong2008@qq.com","fbcdcwapmjtkbicd")


table_format=f'''
<html>
<head></head>
<body>
<table border ="1px solid black">
    <thead>
        <th>name</th>
        <th>sex</sex>
        <th>age</age>
    </thread>
    <tr>
        <td>name</td>
        <td>sex</sex>
        <td>age</age>
    </tr>
    <tr>
        <td>name</td>
        <td>sex</sex>
        <td>age</age>
    </tr>
 </table>
 </body>
 </html>
'''
#for row in sheet:
count = 0
table_col_html = "<thead>" #表头
for row in sheet.iter_rows(min_row=3):
    count +=1
    if count ==1: #first row
       for col in row:
           table_col_html += f"<th>{col.value}</th>"  
       table_col_html += "</thead>"
       continue
    else:    
        row_text="<tr>" #开始一行
        for cell in row:
            #print(cell.value,end=",")
            row_text += f"<td>{cell.value}</td>"
        row_text += "</tr>" # 结束一行

        name = row[2]    
        staff_email = row[1].value
        print(staff_email,name.value)

    mail_body_context=f'''
        <h3>{name.value},你好：</h3>
        <p>请查收你2022-05月的工资条...</p>
        <table border ="1px solid black">
        {table_col_html}
        {row_text}
        </table>
    '''
    msg=MIMEText(mail_body_context,"html","utf-8")

    msg["From"]=Header("Huse人事部","utf-8")# 发送者
    msg["To"] = Header("Huse员工","utf-8")#接收者
    msg["Subject"]=Header("Huse集团2022-05月工资","utf-8")#主题

    #send 
    smtp_obj.sendmail("hnchenhairong2008@qq.com",[staff_email],msg.as_string())
    print(f"成功发送工资条到{staff_email}-{name.value}...")